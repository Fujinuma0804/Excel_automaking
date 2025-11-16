"""
Excel集計レポート生成ツール
「取得データ」「配点」「Template」シートから受講者ごとの集計レポートを生成
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, Reference, Series
import os
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import traceback
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule


class ExcelReportGenerator:
    def __init__(self):
        self.wb = None
        self.data_sheet = None
        self.point_sheet = None
        self.template_sheet = None
        self.original_file_path = None  # 元のファイルパスを保存
        
    def load_workbook(self, file_path):
        """Excelファイルを読み込む"""
        try:
            self.wb = load_workbook(file_path, keep_vba=True)
            self.original_file_path = file_path  # 元のファイルパスを保存
            return True
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みに失敗しました: {str(e)}")
    
    def find_sheets(self):
        """必要なシートを検索"""
        sheet_names = self.wb.sheetnames
        
        # 「取得データ」シートを検索
        for name in sheet_names:
            if '取得データ' in name or '取得' in name:
                self.data_sheet = self.wb[name]
                break
        
        # 「配点」シートを検索
        for name in sheet_names:
            if '配点' in name:
                self.point_sheet = self.wb[name]
                break
        
        # 「Template」シートを検索
        for name in sheet_names:
            if 'template' in name.lower() or 'ひな型' in name or '雛型' in name:
                self.template_sheet = self.wb[name]
                break
        
        if not self.data_sheet:
            raise Exception("「取得データ」シートが見つかりません")
        if not self.point_sheet:
            raise Exception("「配点」シートが見つかりません")
        if not self.template_sheet:
            raise Exception("「Template」シートが見つかりません")
    
    def get_answer_column(self):
        """回答列（Q列）を取得"""
        return 17  # Q列は17列目
    
    def read_point_data(self):
        """配点シートから配点データを読み込む＋分類・設問・総数取得"""
        # 配点シートの構造:
        # B列: セクション名
        # C列: 設問文
        # D列: 問題番号
        # E列: 配点
        points = []
        sections = {}  # セクション別の情報
        section_names = []
        problems = []
        row = 3  # ヘッダー行は2行目、データは3行目から

        while row <= self.point_sheet.max_row:
            # 問題番号を取得（D列）
            question_num_cell = self.point_sheet.cell(row, 4)  # D列
            if question_num_cell.value is None:
                row += 1
                continue

            # セクション名を取得（B列）
            section_cell = self.point_sheet.cell(row, 2)  # B列
            section_name = str(section_cell.value).strip() if section_cell.value else ""
            if section_name and section_name not in section_names:
                section_names.append(section_name)

            # 設問文を取得（C列）
            problem_cell = self.point_sheet.cell(row, 3)  # C列
            problem_text = str(problem_cell.value).strip() if problem_cell.value else ""
            problems.append(problem_text)

            # 配点を取得（E列）
            point_cell = self.point_sheet.cell(row, 5)  # E列
            if point_cell.value is not None:
                question_num = int(question_num_cell.value)
                point_value = float(point_cell.value)

                points.append({
                    'question_num': question_num,
                    'section': section_name,
                    'point': point_value,
                    'problem': problem_text
                })

                # セクション別の集計用
                if section_name not in sections:
                    sections[section_name] = {'total_points': 0, 'questions': []}
                sections[section_name]['total_points'] += point_value
                sections[section_name]['questions'].append(question_num)

            row += 1

        # 問題番号順にソート
        points.sort(key=lambda x: x['question_num'])

        total_problems = len(points)
        return points, sections, section_names, problems, total_problems
    
    
    def read_student_data(self, points_data, sections_data):
        """取得データシートから受講者データを読み込む（各問題ごとの得点を計算）"""
        # 取得データシートの構造:
        # L列（12列目）: 氏名
        # M列（13列目）: メールアドレス
        # Q列（17列目）から: 回答データ（0=不正解、1=正解）
        students = []
        answer_col = self.get_answer_column()  # Q列 = 17列目

        # ヘッダー行は1行目、データは2行目から
        data_start_row = 2

        for row in range(data_start_row, self.data_sheet.max_row + 1):
            # 氏名を取得（L列 = 12列目）
            name_cell = self.data_sheet.cell(row, 12)
            if name_cell.value is None:
                continue

            # 回答データを取得（Q列から）
            answers = []
            col = answer_col  # Q列 = 17列目
            while col <= min(195, self.data_sheet.max_column):
                # 回答は answer_col または answer_col+1 にある場合がある
                answer_cell_1 = self.data_sheet.cell(row, col)
                answer_cell_2 = self.data_sheet.cell(row, col + 1)

                # どちらかに値があれば優先して取得
                answer_value = None
                if answer_cell_1.value is not None and str(answer_cell_1.value).strip() != '':
                    try:
                        answer_value = int(answer_cell_1.value)
                    except (ValueError, TypeError):
                        answer_value = 0
                elif answer_cell_2.value is not None and str(answer_cell_2.value).strip() != '':
                    try:
                        answer_value = int(answer_cell_2.value)
                    except (ValueError, TypeError):
                        answer_value = 0
                else:
                    answer_value = 0

                # 0か1のみ許容
                if answer_value in [0, 1]:
                    answers.append(answer_value)
                else:
                    answers.append(0)

                col += 3

            # 各セクションの得点を計算
            section_scores = {section: 0 for section in sections_data.keys()}
            total_score = 0

            for i, point_info in enumerate(points_data):
                section_name = point_info['section']
                point_value = point_info['point']
                answer = answers[i] if i < len(answers) else 0
                if answer == 1:
                    section_scores[section_name] += point_value
                    total_score += point_value

            students.append({
                'name': str(name_cell.value).strip(),
                'section_scores': section_scores,
                'total_score': total_score,
                'answers': answers,  # ← これを追加
                'row': row
            })

        return students
    
    def calculate_scores(self, students, points_data, sections_data):
        """各受講者の得点を計算"""
        # points_data: 問題番号順にソートされた配点データのリスト
        # sections_data: セクション別の情報
        results = []
        
        for student in students:
            answers = student['answers']
            total_score = 0
            max_score = 0
            section_scores = {}  # セクション別の得点と正解数
            
            # セクション別の集計を初期化
            for section_name in sections_data.keys():
                section_scores[section_name] = {
                    'score': 0,  # 配点を考慮した得点
                    'max_score': sections_data[section_name]['total_points'],  # セクションの満点
                    'correct_count': 0,  # 正解した問題数
                    'total_questions': 0  # セクションの問題数
                }
            
            # セクション別の問題数をカウント
            for point_info in points_data:
                section_name = point_info['section']
                if section_name in section_scores:
                    section_scores[section_name]['total_questions'] += 1
            
            # 配点データと回答を照合
            question_scores = []
            
            for i, point_info in enumerate(points_data):
                question_num = point_info['question_num']
                section_name = point_info['section']
                point_value = point_info['point']
                
                max_score += point_value
                
                # 回答を取得（問題番号は1から始まるので、インデックスはquestion_num-1）
                # 回答データはQ列から3列おきに取得されているため、問題番号順に対応
                if question_num - 1 < len(answers):
                    answer = answers[question_num - 1]
                else:
                    answer = 0  # 回答がない場合は0（不正解）
                
                if answer == 1:  # 正解
                    total_score += point_value
                    section_scores[section_name]['score'] += point_value
                    section_scores[section_name]['correct_count'] += 1
                    question_scores.append({
                        'question_num': question_num,
                        'section': section_name,
                        'point': point_value,
                        'correct': True
                    })
                else:  # 不正解
                    question_scores.append({
                        'question_num': question_num,
                        'section': section_name,
                        'point': point_value,
                        'correct': False
                    })
            
            # 5点評価を計算（100点満点として）
            if max_score > 0:
                percentage = (total_score / max_score) * 100
                if percentage >= 90:
                    rating = 5
                elif percentage >= 80:
                    rating = 4
                elif percentage >= 70:
                    rating = 3
                elif percentage >= 60:
                    rating = 2
                else:
                    rating = 1
            else:
                rating = 0
                percentage = 0
            
            results.append({
                'name': student['name'],
                'total_score': total_score,
                'max_score': max_score,
                'percentage': percentage,
                'rating': rating,
                'section_scores': section_scores,
                'question_scores': question_scores,
                'answers': answers
            })
        
        return results
    
    def calculate_company_averages(self, results, sections_data):
        """全受講者の平均値を計算（5点評価）"""
        company_avg = {}
        section_names = list(sections_data.keys())
        
        for section_name in section_names:
            total_rating = 0
            count = 0
            for result in results:
                section_score = result['section_scores'].get(section_name, {'score': 0, 'max_score': 1})
                if section_score['max_score'] > 0:
                    rating = (section_score['score'] / section_score['max_score']) * 5
                    total_rating += rating
                    count += 1
            company_avg[section_name] = round(total_rating / count, 2) if count > 0 else 0
        
        return company_avg
    
    def create_radar_chart(self, sheet, section_names, data_start_row=27, chart_position="B8"):
        """テーブルデータからレーダーチャートを作成"""
        try:
            # レーダーチャートを作成
            chart = RadarChart()
            chart.type = "standard"  # RadarChartのタイプ: 'standard', 'filled', 'marker'のいずれか
            chart.style = 26
            
            # Y軸のスケール設定（0-5の範囲）
            if hasattr(chart, 'y_axis') and hasattr(chart.y_axis, 'scaling'):
                chart.y_axis.scaling.min = 0
                chart.y_axis.scaling.max = 5
            
            # カテゴリ（ラベル）の参照 - B列のセクション名（行27-31）
            categories = Reference(sheet, min_col=2, min_row=data_start_row, max_row=data_start_row + len(section_names) - 1)
            
            # 社内平均のデータ系列 - C列（行27-31）
            avg_data = Reference(sheet, min_col=3, min_row=data_start_row, max_row=data_start_row + len(section_names) - 1)
            avg_series = Series(avg_data, title="社内平均")
            chart.series.append(avg_series)
            
            # 今回の得点のデータ系列 - D列（行27-31）
            current_data = Reference(sheet, min_col=4, min_row=data_start_row, max_row=data_start_row + len(section_names) - 1)
            current_series = Series(current_data, title="今回の得点")
            chart.series.append(current_series)
            
            # カテゴリを設定
            chart.set_categories(categories)
            
            # チャートをシートに追加（B8セル付近に配置）
            sheet.add_chart(chart, chart_position)
            
        except Exception as e:
            # チャート作成に失敗しても処理を続行
            print(f"チャート作成エラー: {str(e)}")
            import traceback
            traceback.print_exc()
            pass
    
    def create_report_sheet(self, result, template_sheet_name, student_row_index, all_results=None, sections_data=None):
        """
        個別レポートシートを作成（テンプレートをそのままコピーし、氏名と得点のみ埋める）
        グラフや星などテンプレートの内容はそのまま残す
        """
        template = self.wb[template_sheet_name]
        # シート名を学生の名前に設定（Excelのシート名は31文字まで）
        new_sheet_name = result['name']
        if len(new_sheet_name) > 31:
            new_sheet_name = new_sheet_name[:31]
        # 既存のシートがあれば削除
        if new_sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[new_sheet_name])
        # テンプレートシートをコピー（チャートや星レビューも含む）
        # copy_worksheet はチャート、図形、VBAマクロ、条件付き書式などを含むすべての要素をコピーします
        new_sheet = self.wb.copy_worksheet(template)
        new_sheet.title = new_sheet_name
        
        # 重要: 星レビューの視覚的表示には、テンプレートシートに以下の設定が必要です：
        # 1. セルF4:J4を選択
        # 2. 「ホーム」タブ → 「条件付き書式」 → 「アイコンセット」 → 「5つ星の評価」を選択
        # 3. または、カスタム数値書式で星記号（★）を使用
        # これらの書式は copy_worksheet により自動的にコピーされます

        # 氏名をA2セルに記入
        try:
            new_sheet['A2'] = result['name']
        except Exception:
            pass

        

        # スキルカテゴリごとの得点（D27～D31）を埋める
        section_names = list(result['section_scores'].keys())
        section_row_mapping = {
            0: 27,
            1: 28,
            2: 29,
            3: 30,
            4: 31,
        }
        
        # 社内平均を計算（全受講者の平均値）
        company_avg = {}
        if all_results and sections_data:
            company_avg = self.calculate_company_averages(all_results, sections_data)
        
        for idx, section_name in enumerate(section_names):
            if idx in section_row_mapping:
                row = section_row_mapping[idx]
                # セクション名をB列に設定
                try:
                    new_sheet.cell(row, 2).value = section_name
                except Exception:
                    pass
                
                # 社内平均をC列に設定
                if section_name in company_avg:
                    try:
                        new_sheet.cell(row, 3).value = company_avg[section_name]
                    except Exception:
                        pass
                
                # 5点評価を計算してD列に設定
                section_score = result['section_scores'].get(section_name, {'score': 0, 'max_score': 1})

                if section_score['max_score'] > 0:
                    section_value = round((section_score['score'] / section_score['max_score']) * 5, 2)
                else:
                    section_value = 0
                try:
                    new_sheet.cell(row, 4).value = section_value
                except Exception:
                    pass

        # 総合点（E4セル）を埋める（例: 5点評価の平均値）
        try:
            avg_rating = round(
                sum(
                    round((section['score'] / section['max_score']) * 5, 2)
                    if section['max_score'] > 0 else 0
                    for section in result['section_scores'].values()
                ) / len(result['section_scores']), 2
            ) if len(result['section_scores']) > 0 else 0
            new_sheet['E4'] = avg_rating
            
            # 星レビュー用の数式を設定（F4～J4）
            # E4の値に基づいて、各星セルに★（満点）、半分の星（左半分が塗りつぶし）、または☆（空）を表示
            # 5.0 = ★★★★★, 4.0-4.9 = ★★★★◐ (4 full + 1 half), 3.0-3.9 = ★★★◐☆, etc.
            # 半分の星は左半分が塗りつぶされた星を表現（視覚的に半分に見えるように）
            try:
                # 星の表示ロジック：
                # - 整数部分の星は★（満点、塗りつぶし）
                # - 小数部分がある場合、次の星を半分として表示（左半分塗りつぶしの星）
                # - 半分の星はUnicodeの左半分円（◐）または視覚的に半分に見える文字を使用
                
                # F4: 1つ目の星（E4>=1なら★、それ以外は☆）
                new_sheet['F4'] = '=IF(E4>=1,"★","☆")'
                
                # G4: 2つ目の星（E4>=2なら★、E4>=1かつE4<2なら半分★、それ以外は☆）
                # 半分の星には左半分が塗りつぶされた円（◐）または視覚的に半分に見える文字を使用
                # 実際には、条件付き書式で半分の星を表示する方が良いが、数式でも可能
                new_sheet['G4'] = '=IF(E4>=2,"★",IF(AND(E4>=1,E4<2),"◐","☆"))'
                
                # H4: 3つ目の星（E4>=3なら★、E4>=2かつE4<3なら半分★、それ以外は☆）
                new_sheet['H4'] = '=IF(E4>=3,"★",IF(AND(E4>=2,E4<3),"◐","☆"))'
                
                # I4: 4つ目の星（E4>=4なら★、E4>=3かつE4<4なら半分★、それ以外は☆）
                new_sheet['I4'] = '=IF(E4>=4,"★",IF(AND(E4>=3,E4<4),"◐","☆"))'
                
                # J4: 5つ目の星（E4>=5なら★、E4>=4かつE4<5なら半分★、それ以外は☆）
                new_sheet['J4'] = '=IF(E4>=5,"★",IF(AND(E4>=4,E4<5),"◐","☆"))'
                
                # フォントスタイルを設定して星を見やすくする
                star_cells = ['F4', 'G4', 'H4', 'I4', 'J4']
                try:
                    for cell_ref in star_cells:
                        cell = new_sheet[cell_ref]
                        # 星の色を金色（FFD700）に設定し、サイズを大きく
                        # 半分の星（◐）も同じスタイルで表示される
                        cell.font = Font(name='Arial', size=16, color='FFD700', bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception:
                    pass
                
                # 注意: ◐（左半分円）は星の形ではないため、視覚的に最適ではない可能性があります
                # Excelで半分の星を正確に表示するには、条件付き書式のアイコンセットまたは
                # カスタム図形を使用する方が良いですが、数式ベースの実装として◐を使用します
                
            except Exception as e:
                print(f"星レビュー数式設定エラー: {str(e)}")
                import traceback
                traceback.print_exc()
                pass
                
        except Exception:
            pass

        # レーダーチャートを作成（テーブルデータを基に）
        try:
            self.create_radar_chart(new_sheet, section_names, data_start_row=27, chart_position="B8")
        except Exception as e:
            # チャート作成に失敗しても処理を続行
            print(f"チャート作成エラー: {str(e)}")
            import traceback
            traceback.print_exc()
            pass

        # 必要ならコメント欄なども result から埋める
        # 例: new_sheet['B35'] = result.get('comment', '')

        return new_sheet
    
    def create_summary_sheet(self, results, sections_data, points_data):
        """集計シートを作成（分類別得点を正確に集計）"""
        summary_name = "総合得点"
        if summary_name in self.wb.sheetnames:
            self.wb.remove(self.wb[summary_name])
        summary_sheet = self.wb.create_sheet(summary_name)

        # ヘッダー行
        headers = ['氏名'] + list(sections_data.keys()) + ['総合得点（満点）']
        header_colors = [
            "B7DEE8", "DCE6F1", "FDE9D9", "EAF1DD", "E4DFEC", "FDE9D9", "F8CBAD",
        ]
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(2, col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            if col-1 < len(header_colors):
                fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor=header_colors[col-1]
                )
                cell.fill = fill

        section_question_map = {section: [] for section in sections_data.keys()}
        for pt in points_data:
            section_question_map[pt['section']].append(pt['question_num'])

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # データ行
        for row_idx, result in enumerate(results, 3):
            summary_sheet.cell(row_idx, 1).value = result['name']
            summary_sheet.cell(row_idx, 1).border = thin_border
            total_score = 0
            col_idx = 2
            for section_name in sections_data.keys():
                question_nums = section_question_map[section_name]
                section_score = 0
                for qn in question_nums:
                    if qn - 1 < len(result['answers']):
                        answer = result['answers'][qn - 1]
                        if answer == 1:
                            for pt in points_data:
                                if pt['question_num'] == qn:
                                    section_score += pt['point']
                                    break
                cell = summary_sheet.cell(row_idx, col_idx)
                cell.value = section_score
                cell.border = thin_border
                total_score += section_score
                col_idx += 1
            max_score = sum([pt['point'] for pt in points_data])
            total_cell = summary_sheet.cell(row_idx, col_idx)
            total_cell.value = int(total_score)
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal='right')

        # 平均行の追加
        avg_row_idx = len(results) + 3
        summary_sheet.cell(avg_row_idx, 1).value = "平均"
        summary_sheet.cell(avg_row_idx, 1).font = openpyxl.styles.Font(bold=True)
        summary_sheet.cell(avg_row_idx, 1).border = thin_border

        for col in range(2, len(headers) + 1):
            col_letter = get_column_letter(col)
            # 3行目からデータが始まる
            formula = f"=AVERAGE({col_letter}3:{col_letter}{avg_row_idx-1})"
            cell = summary_sheet.cell(avg_row_idx, col)
            cell.value = formula
            cell.border = thin_border
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'

        # ヘッダー行にも枠線を追加
        for col in range(1, len(headers) + 1):
            cell = summary_sheet.cell(2, col)
            cell.border = thin_border
            summary_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    def create_rating_sheet(self, results, sections_data):
        """5点評価シートを作成（各セクションごとに5点評価を表示・集計）"""
        rating_name = "5点評価"
        if rating_name in self.wb.sheetnames:
            self.wb.remove(self.wb[rating_name])
        rating_sheet = self.wb.create_sheet(rating_name)

        # ヘッダー行
        headers = ['氏名'] + list(sections_data.keys()) + ['総合評価（5点満点）']
        header_colors = [
            "B7DEE8", "DCE6F1", "FDE9D9", "EAF1DD", "E4DFEC", "FDE9D9", "F8CBAD",
        ]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = rating_sheet.cell(2, col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = thin_border
            if col-1 < len(header_colors):
                fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor=header_colors[col-1]
                )
                cell.fill = fill
            rating_sheet.column_dimensions[get_column_letter(col)].width = 20

        # データ行
        for row_idx, result in enumerate(results, 3):
            rating_sheet.cell(row_idx, 1).value = result['name']
            rating_sheet.cell(row_idx, 1).border = thin_border
            total_rating = 0
            col_idx = 2
            for section_name in sections_data.keys():
                # 5点評価を取得
                section_score = result['section_scores'].get(section_name, {'score': 0, 'max_score': 1})
                if section_score['max_score'] > 0:
                    section_value = (section_score['score'] / section_score['max_score']) * 5                    
                else:
                    section_value = 0
                cell = rating_sheet.cell(row_idx, col_idx)
                cell.value = section_value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '0.00'
                total_rating += section_value
                col_idx += 1
            # 総合評価（5点満点）の平均
            avg_rating = round(total_rating / len(sections_data), 2) if len(sections_data) > 0 else 0
            total_cell = rating_sheet.cell(row_idx, col_idx)
            total_cell.value = avg_rating
            total_cell.border = thin_border
            total_cell.number_format = '0.00'
            total_cell.alignment = Alignment(horizontal='right')

        # 平均行の追加
        avg_row_idx = len(results) + 3
        rating_sheet.cell(avg_row_idx, 1).value = "平均"
        rating_sheet.cell(avg_row_idx, 1).font = openpyxl.styles.Font(bold=True)
        rating_sheet.cell(avg_row_idx, 1).border = thin_border

        for col in range(2, len(headers) + 1):
            col_letter = get_column_letter(col)
            formula = f"=AVERAGE({col_letter}3:{col_letter}{avg_row_idx-1})"
            cell = rating_sheet.cell(avg_row_idx, col)
            cell.value = formula
            cell.border = thin_border
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'
    
    def update_data_sheet(self, students, results, sections_data):
        """取得データシートに各問題類型のスコア列を追加"""
        # 学生の行番号と結果を対応付ける辞書を作成
        student_row_to_result = {}
        for student, result in zip(students, results):
            student_row_to_result[student['row']] = result
        
        # 最後の列を取得
        last_col = self.data_sheet.max_column
        start_col = last_col + 2  # 1列空けてから追加
        
        # ヘッダー行（1行目）に各問題類型の名前を追加
        header_row = 1
        section_names = list(sections_data.keys())
        for col_idx, section_name in enumerate(section_names):
            col = start_col + col_idx
            cell = self.data_sheet.cell(header_row, col)
            cell.value = section_name
            cell.font = openpyxl.styles.Font(bold=True)
        
        # 各学生の行にスコアを追加
        for row_num in range(2, self.data_sheet.max_row + 1):
            if row_num in student_row_to_result:
                result = student_row_to_result[row_num]
                for col_idx, section_name in enumerate(section_names):
                    col = start_col + col_idx
                    section_score = result['section_scores'].get(section_name, {'score': 0})
                    cell = self.data_sheet.cell(row_num, col)
                    cell.value = section_score['score']
    
    def generate_reports(self, output_path=None):
        """レポートを生成"""
        try:
            # シートを検索
            self.find_sheets()
            
            # データを読み込む
            points_data, sections_data, section_names, problems, total_problems = self.read_point_data()
            students = self.read_student_data(points_data, sections_data)
            
            if not students:
                raise Exception("受講者データが見つかりません")
            
            if not points_data:
                raise Exception("配点データが見つかりません")
            
            # 得点を計算
            results = self.calculate_scores(students, points_data, sections_data)
            
            # 取得データシートに各問題類型のスコア列を追加
            self.update_data_sheet(students, results, sections_data)
            
            # 集計シートを作成
            self.create_summary_sheet(results, sections_data, points_data)
            
            # 5点評価シートを作成
            self.create_rating_sheet(results, sections_data)
            
            # 個別レポートシートを作成
            template_sheet_name = self.template_sheet.title
            for idx, result in enumerate(results, 3):  # 3行目から開始（ヘッダー行が2行目）
                self.create_report_sheet(result, template_sheet_name, idx, all_results=results, sections_data=sections_data)
            
            # ファイルを保存
            if output_path:
                # 出力パスが指定されている場合
                base_output_path = Path(output_path)
            else:
                # 元のファイル名に「_出力」を追加
                if self.original_file_path:
                    base_path = Path(self.original_file_path)
                    base_output_path = base_path.parent / f"{base_path.stem}_出力{base_path.suffix}"
                else:
                    raise Exception("元のファイルパスが設定されていません")
            
            # 既存のファイルが存在し、開かれている場合はタイムスタンプを追加
            output_path_obj = base_output_path
            if output_path_obj.exists():
                try:
                    # 削除を試みる（開かれていない場合）
                    output_path_obj.unlink()
                    output_path_str = str(output_path_obj)
                except (PermissionError, OSError):
                    # ファイルが開かれている場合はタイムスタンプを追加
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    base_name = output_path_obj.stem
                    suffix = output_path_obj.suffix
                    output_path_str = str(output_path_obj.parent / f"{base_name}_{timestamp}{suffix}")
            else:
                # ファイルが存在しない場合はそのまま使用
                output_path_str = str(output_path_obj)
            
            # ファイルを保存
            try:
                self.wb.save(output_path_str)
            except PermissionError:
                raise Exception(
                    f"ファイルの保存に失敗しました。\n"
                    f"以下の可能性があります：\n"
                    f"1. 出力ファイルが既にExcelで開かれている\n"
                    f"2. ファイルのアクセス権限がない\n"
                    f"3. ディレクトリへの書き込み権限がない\n\n"
                    f"ファイル: {Path(output_path_str).name}\n"
                    f"パス: {Path(output_path_str).parent}"
                )
            
            return results, output_path_str
            
        except Exception as e:
            raise Exception(f"レポート生成中にエラーが発生しました: {str(e)}\n{traceback.format_exc()}")


class ReportGeneratorUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel集計レポート生成ツール")
        self.root.geometry("600x400")
        
        self.generator = ExcelReportGenerator()
        self.file_path = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIを構築"""
        # メインフレーム
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # ファイル選択
        file_frame = ttk.LabelFrame(main_frame, text="Excelファイル選択", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.file_label = ttk.Label(file_frame, text="ファイルが選択されていません", foreground="gray")
        self.file_label.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(file_frame, text="ファイルを選択", command=self.select_file).pack(side=tk.RIGHT)
        
        # 実行ボタン
        execute_frame = ttk.Frame(main_frame)
        execute_frame.pack(fill=tk.X, pady=10)
        
        self.execute_button = ttk.Button(
            execute_frame,
            text="レポートを生成",
            command=self.generate_reports,
            state=tk.DISABLED
        )
        self.execute_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 進捗表示
        self.progress = ttk.Progressbar(execute_frame, mode='indeterminate')
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # ログ表示
        log_frame = ttk.LabelFrame(main_frame, text="ログ", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.log_text = tk.Text(log_frame, height=15, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log("ツールを起動しました。")
        self.log("Excelファイルを選択してください。")
    
    def log(self, message):
        """ログを追加"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def select_file(self):
        """ファイルを選択"""
        file_path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        
        if (file_path):
            self.file_path = file_path
            self.file_label.config(text=os.path.basename(file_path), foreground="black")
            self.execute_button.config(state=tk.NORMAL)
            self.log(f"ファイルを選択しました: {os.path.basename(file_path)}")
    
    def generate_reports(self):
        """レポートを生成"""
        if not self.file_path:
            messagebox.showerror("エラー", "ファイルを選択してください。")
            return
        
        try:
            self.execute_button.config(state=tk.DISABLED)
            self.progress.start()
            self.log("Excelファイルを読み込んでいます...")
            
            # ファイルを読み込む
            self.generator.load_workbook(self.file_path)
            self.log("ファイルの読み込みが完了しました。")
            
            # レポートを生成
            self.log("レポートを生成しています...")
            results, output_path = self.generator.generate_reports()
            
            self.progress.stop()
            self.execute_button.config(state=tk.NORMAL)
            
            self.log(f"レポート生成が完了しました！")
            self.log(f"出力ファイル: {output_path}")
            self.log(f"処理した受講者数: {len(results)}名")
            
            messagebox.showinfo(
                "完了",
                f"レポート生成が完了しました。\n\n"
                f"出力ファイル: {os.path.basename(output_path)}\n"
                f"処理した受講者数: {len(results)}名"
            )
            
        except Exception as e:
            self.progress.stop()
            self.execute_button.config(state=tk.NORMAL)
            error_msg = str(e)
            self.log(f"エラー: {error_msg}")
            messagebox.showerror("エラー", f"レポート生成中にエラーが発生しました:\n{error_msg}")


def main():
    root = tk.Tk()
    app = ReportGeneratorUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

